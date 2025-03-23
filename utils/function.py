from datetime import datetime
import pytz
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from collections import defaultdict
import utils.calculate as calculate


def check_type_expense_insurance(name):
    if name == "life_insurance":
        return "ประกันชีวิต"
    if name == "car_insurance":
        return "ประกันรถยนต์"
    if name == "home_insurance":
        return "ประกันบ้าน"
    if name == "social security":
        return "ประกันสังคม"
    if name == "Provident fund savings":
        return "เงินสะสมกองทุนสำรองเลี้ยงชีพ"
    
def check_type_expense_var(name):
    if name == "food_expense":
        return "ค่าอาหาร"
    if name == "utility_expense":
        return "ค่าใช้จ่ายสาธารณูปโภค"
    if name == "telephone_charge":
        return "ค่าโทรศัพท์"
    if name == "medical_expense":
        return "ค่ารักษาพยาบาล"
    if name == "Monthly expenses for family members or pets":
        return "ค่าใช้จ่ายให้คนในครอบครัวหรือสัตว์เลี้ยงต่อเดือน"
    if name == "Other expenses":
        return "ค่าใช้จ่ายอื่น ๆ"

def check_type_long_term(name):
    if name == "mortgage-loans":
        return "สินเชื่อเพื่อที่อยู่อาศัย"
    if name == "education-loans":
        return "สินเชื่อเพื่อการศึกษา"
    if name == "car-loans":
        return "สินเชื่อเพื่อยานพาหนะ"
    if name == "credit-card-long-term-debt":
        return "หนี้สินจากบัตรเครดิตระยะยาว"
    if name == "business-loans":
        return "หนี้สินเพื่อการลงทุน"
    if name == "personal-long-term-loans":
        return "หนี้สินเพื่อการบริโภคระยะยาว"
    if name == "insurance-loans":
        return "หนี้สินเพื่อการประกันและความมั่นคง"
    if name == "other":
        return "อื่นๆ"

def check_type_short_term(name):
    if name == "credit-card":
        return "หนี้สินจากบัตรเครดิต"
    if name == "short-term-loan":
        return "หนี้สินจากการกู้ยืมระยะสั้น"
    if name == "daily-loan":
        return "หนี้ค้างจ่ายในชีวิตประจำวัน"
    if name == "installment-payments":
        return "หนี้สินจากการผ่อนชำระสินค้า"
    if name == "student-loan":
        return "หนี้สินจากการกู้เพื่อการศึกษา"
    if name == "security-deposits":
        return "หนี้สินจากการวางเงินล่วงหน้า"
    if name == "taxes":
        return "หนี้สินจากการชำระภาษี"
    if name == "other":
        return "อื่นๆ"


def summanry_expense(data):
    sum = 0
    for item in data:
        if item['fix_cf_value'] == "":
            sum = sum + 0
        else:
            sum = sum + float(item['fix_cf_value'])
    
    return sum



def summanry_income(data):
    sum = 0
    if data['income_1'] == "":
        sum = sum + 0
    else:
        sum = sum + float(data[0]['income_1'])

    if data['income_2'] == "":
        sum = sum + 0
    else:
        sum = sum + float(data[0]['income_2'])

    if data['income_3'] == "":
        sum = sum + 0
    else:
        sum = sum + float(data[0]['income_3'])

    if data['income_4'] == "":
        sum = sum + 0
    else:
        sum = sum + float(data[0]['income_4'])

    if data['income_5'] == "":
        sum = sum + 0
    else:
        sum = sum + float(data[0]['income_5'])

    if data['income_6'] == "":
        sum = sum + 0
    else:
        sum = sum + float(data['income_6'])

    if data['income_7'] == "":
        sum = sum + 0
    else:
        sum = sum + float(data['income_7'])

    if data['income_8'] == "":
        sum = sum + 0
    else:
        sum = sum + float(data['income_8'])

    print("sum : " , sum)
    return sum


def calculate_age(birthdate_str):
    thailand_tz = pytz.timezone('Asia/Bangkok')
    birthdate = datetime.strptime(birthdate_str, "%Y-%m-%d").date()
    today = datetime.now(thailand_tz).date()
    age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day)) 
    return age

def calculate_age_new(birthdate_str):
    birthdate_str = birthdate_str.strftime('%Y-%m-%d')
    thailand_tz = pytz.timezone('Asia/Bangkok')
    birthdate = datetime.strptime(birthdate_str, "%Y-%m-%d").date()
    today = datetime.now(thailand_tz).date()
    age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day)) 
    return str(age)

def calculate_investment(principal=None, future_value=None, annual_contribution=None, annual_rate=None, years=None):
    try:
        principal = Decimal(principal) if principal is not None else None
        future_value = Decimal(future_value) if future_value is not None else None
        annual_contribution = Decimal(annual_contribution) if annual_contribution is not None else None
        annual_rate = Decimal(annual_rate)
        years = Decimal(years) if years is not None else None

        if sum([principal is None, future_value is None, annual_contribution is None, years is None]) != 1:
            raise ValueError("You must leave exactly one parameter as None to solve for it.")

        if future_value is None:
            fv_principal = principal * ((1 + (annual_rate / 100)) ** years)
            fv_contributions = Decimal(0)
            if annual_contribution:
                fv_contributions = annual_contribution * (((1 + (annual_rate / 100)) ** years - 1) / (annual_rate / 100))
            return (fv_principal + fv_contributions)

    except (InvalidOperation, TypeError, ValueError) as e:
        raise ValueError(f"Invalid input for calculation: {e}")


def create_cash_flow_excel(asset , asset_detail , income , work_detail , expense_detail , goal_input1 , personal_info ,financial_info , filename="cash_flow.xlsx"):
    row = 1
    wb = Workbook()
    ws = wb.active
    ws.title = "cash flow"
    
    header_fill = PatternFill(start_color="d5a6bd", end_color="d5a6bd", fill_type="solid")  
    subheader_fill = PatternFill(start_color="b4a7d6", end_color="b4a7d6", fill_type="solid") 
    footheader_fill = PatternFill(start_color="b4a7d6", end_color="b4a7d6", fill_type="solid") 
    
    thin_border = Border(left=Side(style='thin', color="000000"),
                         right=Side(style='thin', color="000000"),
                         top=Side(style='thin', color="000000"),
                         bottom=Side(style='thin', color="000000"))
    bold_border = Border(left=Side(style='medium', color="000000"),
                         right=Side(style='medium', color="000000"),
                         top=Side(style='medium', color="000000"),
                         bottom=Side(style='medium', color="000000"))
    
    ws.append(["รายรับ-รายจ่าย", "", ""])  
    ws.merge_cells("A"+str(row)+":A"+str(row+1))
    ws["A"+str(row)].alignment = Alignment(horizontal="center")
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = subheader_fill
    ws["A"+str(row)].font = Font(bold=True)

    
    ws.merge_cells("B"+str(row)+":G"+str(row))
    ws["B"+str(row)] = "ปี"
    ws["B"+str(row)].alignment = Alignment(horizontal="center")
    ws["B"+str(row)].font = Font(bold=True)
    ws["B"+str(row)].fill = subheader_fill
    ws["B"+str(row)].font = Font(bold=True)

    # ws.append(["" , "2567","2568","2569","2570","2571" , "2572"])  
    yaer = ["2567","2568","2569","2570","2571" , "2572"]
    row_list = ["B","C","D","E","F","G"]

    for item in range(len(yaer)): 
        print("row_list[yaer[item]]+2 : " , row_list[item]+"2")
        ws[row_list[item]+"2"] = str(yaer[item])
        ws[row_list[item]+"2"].alignment = Alignment(horizontal="center")
        ws[row_list[item]+"2"].font = Font(bold=True)
        ws[row_list[item]+"2"].fill = subheader_fill
        ws[row_list[item]+"2"].font = Font(bold=True)


    ws["A3"] = "รายรับ"
    ws["A3"].alignment = Alignment(horizontal="center")
    ws["A3"].font = Font(bold=True)
    ws["A3"].fill = header_fill

    ws["B3"].fill = header_fill
    ws["C3"].fill = header_fill
    ws["D3"].fill = header_fill
    ws["E3"].fill = header_fill
    ws["F3"].fill = header_fill
    ws["G3"].fill = header_fill
    
    print("start")

    salary_growth = (float(work_detail['salary_growth']) / 100) + 1
    growth_rate = 0.00
    growth_rate_order = 0
    inflation = 3

    inflation_set = (1 + (inflation/100))

    for item in asset_detail:
        # print("asset_detail : " , item['growth_rate'])
        if item['growth_rate'] != "":
            growth_rate = growth_rate + float(item['growth_rate'])
            growth_rate_order = growth_rate_order + 1

    print("growth_rate", growth_rate)
    print("salary_growth :" , salary_growth)
    # print("growth_rate / growth_rate_order" , (growth_rate / growth_rate_order))
    if growth_rate_order == 0:
        dividend_growth = 0
    else:
        dividend_growth = (growth_rate / growth_rate_order)
    growth_rate_sum = (1 + (dividend_growth/100))

    income_1  =float(income['income_1']) if income['income_1'] != "" else 0
    income_2  =float(income['income_2']) if income['income_2'] != "" else 0
    income_3  =float(income['income_3']) if income['income_3'] != "" else 0
    income_4  =float(income['income_4']) if income['income_4'] != "" else 0
    income_5  =float(income['income_5']) if income['income_5'] != "" else 0
    income_6  =float(income['income_6']) if income['income_6'] != "" else 0
    income_7  =float(income['income_7']) if income['income_7'] != "" else 0
    income_8  =float(income['income_8']) if income['income_8'] != "" else 0

    sumother = income_2 + income_3 + income_5 + income_6 +income_7 + income_8

    print("---- income_1 ", income_1)
    print("growth_rate_sum: " , growth_rate_sum)
    income_1_2568  = ( income_1 * (float(salary_growth)))
    print("------ " , income_1_2568)
    income_1_2569  = (income_1_2568 * float(salary_growth))
    income_1_2570  = (income_1_2569 * float(salary_growth))
    income_1_2571  = (income_1_2570 * float(salary_growth))
    income_1_2572  = (income_1_2571 * float(salary_growth))

    income_4_2568  = (income_4 * float(growth_rate_sum))
    income_4_2569  = (income_4_2568 * float(growth_rate_sum))
    income_4_2570  = (income_4_2569 * float(growth_rate_sum))
    income_4_2571  = (income_4_2570 * float(growth_rate_sum))
    income_4_2572  = (income_4_2571 * float(growth_rate_sum))

    print("sumother :" , sumother)
    data = [
        ["เงินเดือน(รวมโบนัส)" , "{:,.2f}".format(income_1) , "{:,.2f}".format(income_1_2568)  , "{:,.2f}".format(income_1_2569) , "{:,.2f}".format(income_1_2570) , "{:,.2f}".format(income_1_2571) , "{:,.2f}".format(income_1_2572)],
        ["รายรับอื่นๆ" , "{:,.2f}".format(sumother)  , "{:,.2f}".format(sumother) , "{:,.2f}".format(sumother) , "{:,.2f}".format(sumother) , "{:,.2f}".format(sumother) , "{:,.2f}".format(sumother)],
        ["เงินปันผลจากการลงทุน" , "{:,.2f}".format(income_4) , "{:,.2f}".format(income_4_2568) , "{:,.2f}".format(income_4_2569) , "{:,.2f}".format(income_4_2570) , "{:,.2f}".format(income_4_2571) , "{:,.2f}".format(income_4_2571)],
    ]

    for i in data:
        ws.append(i)   

    ws.append(["รายรับรวม" , "{:,.2f}".format(income_1 + sumother + income_4) , "{:,.2f}".format(income_1_2568 + sumother + income_4_2568) , "{:,.2f}".format(income_1_2569 + sumother + income_4_2569) , "{:,.2f}".format(income_1_2570 + sumother + income_4_2570) , "{:,.2f}".format(income_1_2571 + sumother + income_4_2571) , "{:,.2f}".format(income_1_2572 + sumother + income_4_2571)]) 
    ws["B8"].font = Font(color="FF0000") 
    ws["B8"].font = Font(color="FF0000") 
    ws["C8"].font = Font(color="FF0000") 
    ws["D8"].font = Font(color="FF0000") 
    ws["E8"].font = Font(color="FF0000") 
    ws["F8"].font = Font(color="FF0000") 
    ws["G8"].font = Font(color="FF0000") 


    ws["A9"] = "รายจ่าย"
    ws["A9"].alignment = Alignment(horizontal="center")
    ws["A9"].font = Font(bold=True)
    ws["A9"].fill = header_fill
    ws["B9"].fill = header_fill
    ws["C9"].fill = header_fill
    ws["D9"].fill = header_fill
    ws["E9"].fill = header_fill
    ws["F9"].fill = header_fill
    ws["G9"].fill = header_fill

    fix = 0
    insurance = 0
    varSum = 0
    expense_sum = 0
    for item in expense_detail:
        if item['fix_cf_value'] != "":
            expense_sum = expense_sum + int(item['fix_cf_value'])
        if item['fix_cf_mode'] == "fix":
            if item['fix_cf_value'] != "":
                fix = fix +  float(item['fix_cf_value'])
        if item['fix_cf_mode'] == "insurance":
            if item['fix_cf_value'] != "":
                insurance = insurance +  float(item['fix_cf_value'])
        if item['fix_cf_mode'] == "var":
            if item['fix_cf_value'] != "":
                varSum = varSum +  float(item['fix_cf_value'])
        
    
    fix_2568  = (fix * float(inflation_set))
    fix_2569  = (fix_2568 * float(inflation_set))
    fix_2570  = (fix_2569 * float(inflation_set))
    fix_2571  = (fix_2570 * float(inflation_set))
    fix_2572  = (fix_2571 * float(inflation_set))

    insurance_2568  = (insurance * float(inflation_set))
    insurance_2569  = (insurance_2568 * float(inflation_set))
    insurance_2570  = (insurance_2569 * float(inflation_set))
    insurance_2571  = (insurance_2570 * float(inflation_set))
    insurance_2572  = (insurance_2571 * float(inflation_set))
    

    data = [
        ["กระแสเงินสดจ่ายคงที่/เงินงวดผ่อนชำระคืนหนี้" , "{:,.2f}".format(fix) , "{:,.2f}".format(fix_2568) , "{:,.2f}".format(fix_2569) , "{:,.2f}".format(fix_2570) , "{:,.2f}".format(fix_2571) , "{:,.2f}".format(fix_2572)],
        ["เบี้ยประกัน" , "{:,.2f}".format(insurance) , "{:,.2f}".format(insurance_2568) , "{:,.2f}".format(insurance_2569) , "{:,.2f}".format(insurance_2570) , "{:,.2f}".format(insurance_2571) , "{:,.2f}".format(insurance_2572)]
       
    ]

    for i in data:
        ws.append(i) 

    ws.append(["รวมรายจ่ายคงที่" , "{:,.2f}".format(fix + insurance) , "{:,.2f}".format(fix_2568 + insurance_2568) , "{:,.2f}".format(fix_2569 + insurance_2569) , "{:,.2f}".format(fix_2570 + insurance_2570) , "{:,.2f}".format(fix_2571 + insurance_2571) , "{:,.2f}".format(fix_2572 + insurance_2572)]) 
    ws["A12"].font = Font(color="FF0000") 
    ws["B12"].font = Font(color="FF0000") 
    ws["C12"].font = Font(color="FF0000") 
    ws["D12"].font = Font(color="FF0000") 
    ws["E12"].font = Font(color="FF0000") 
    ws["F12"].font = Font(color="FF0000") 
    ws["G12"].font = Font(color="FF0000") 


    varSum_2568  = (varSum * float(inflation_set))
    varSum_2569  = (varSum_2568 * float(inflation_set))
    varSum_2570  = (varSum_2569 * float(inflation_set))
    varSum_2571  = (varSum_2570 * float(inflation_set))
    varSum_2572  = (varSum_2571 * float(inflation_set))

    data = [
        ["ค่าใช้จ่ายผันแปร" , "{:,.2f}".format(varSum) , "{:,.2f}".format(varSum_2568)  , "{:,.2f}".format(varSum_2569)  , "{:,.2f}".format(varSum_2570)  , "{:,.2f}".format(varSum_2571)  , "{:,.2f}".format(varSum_2572) ]
    ]

    for i in data:
        ws.append(i) 

    ws.append(["รวมรายจ่าย" , "{:,.2f}".format((fix + insurance) + varSum) , "{:,.2f}".format((fix_2568 + insurance_2568) + varSum_2568) , "{:,.2f}".format((fix_2569 + insurance_2569) + varSum_2569) , "{:,.2f}".format((fix_2570 + insurance_2570) + varSum_2570) , "{:,.2f}".format((fix_2571 + insurance_2571) + varSum_2571) , "{:,.2f}".format((fix_2572 + insurance_2572) + varSum_2572)]) 
    ws["A14"].font = Font(color="FF0000") 
    ws["B14"].font = Font(color="FF0000") 
    ws["C14"].font = Font(color="FF0000") 
    ws["D14"].font = Font(color="FF0000") 
    ws["E14"].font = Font(color="FF0000") 
    ws["F14"].font = Font(color="FF0000") 
    ws["G14"].font = Font(color="FF0000") 

    ws.append(["กระแสเงินสุทธิก่อนเป้าหมาย" , "{:,.2f}".format( (income_1 + sumother + income_4) - ((fix + insurance) + varSum)) , "{:,.2f}".format( (income_1_2568 + sumother + income_4_2568) - ((fix_2568 + insurance_2568) + varSum_2568)) , "{:,.2f}".format( (income_1_2569 + sumother + income_4_2569) - ((fix_2569 + insurance_2569) + varSum_2569)) , "{:,.2f}".format( (income_1_2570 + sumother + income_4_2570) - ((fix_2570 + insurance_2570) + varSum_2570)) , "{:,.2f}".format( (income_1_2571 + sumother + income_4_2571) - ((fix_2571 + insurance_2571) + varSum_2571)) , "{:,.2f}".format( (income_1_2572 + sumother + income_4_2572) - ((fix_2572 + insurance_2572) + varSum_2572))]) 
    ws["A15"].font = Font(color="FF0000") 
    ws["B15"].font = Font(color="FF0000") 
    ws["C15"].font = Font(color="FF0000") 
    ws["D15"].font = Font(color="FF0000") 
    ws["E15"].font = Font(color="FF0000") 
    ws["F15"].font = Font(color="FF0000") 
    ws["G15"].font = Font(color="FF0000")

    sum_all = (income_1 + sumother + income_4) - ((fix + insurance) + varSum)
    sum_all_2668 = sum_all + ((income_1_2568 + sumother + income_4_2568) - ((fix_2568 + insurance_2568) + varSum_2568))
    sum_all_2669 = sum_all_2668 + ((income_1_2569 + sumother + income_4_2569) - ((fix_2569 + insurance_2569) + varSum_2569))
    sum_all_2670 = sum_all_2669 + ((income_1_2570 + sumother + income_4_2570) - ((fix_2570 + insurance_2570) + varSum_2570))
    sum_all_2671 = sum_all_2670 + ((income_1_2571 + sumother + income_4_2571) - ((fix_2571 + insurance_2571) + varSum_2571))
    sum_all_2672 = sum_all_2671 + ((income_1_2572 + sumother + income_4_2572) - ((fix_2572 + insurance_2572) + varSum_2570))

    ws.append(["กระแสเงินสุทธิทบยอด" , "{:,.2f}".format(sum_all) , "{:,.2f}".format(sum_all_2668) , "{:,.2f}".format(sum_all_2669) , "{:,.2f}".format(sum_all_2670) , "{:,.2f}".format(sum_all_2671) , "{:,.2f}".format(sum_all_2672)]) 
    # ws["A16"].font = Font(color="FF0000") 
    # ws["B16"].font = Font(color="FF0000") 
    # ws["C16"].font = Font(color="FF0000") 
    # ws["D16"].font = Font(color="FF0000") 
    # ws["E16"].font = Font(color="FF0000") 
    # ws["F16"].font = Font(color="FF0000") 
    # ws["G16"].font = Font(color="FF0000")


    ws["A17"] = "รายจ่ายเพื่อเป้าหมายทางการเงิน"
    ws["A17"].alignment = Alignment(horizontal="center")
    ws["A17"].font = Font(bold=True)
    ws["A17"].fill = header_fill
    ws["B17"].fill = header_fill
    ws["C17"].fill = header_fill
    ws["D17"].fill = header_fill
    ws["E17"].fill = header_fill
    ws["F17"].fill = header_fill
    ws["G17"].fill = header_fill

    row = 17
    data_goal = []
    for item in goal_input1:
        print("item : " ,item)
        if item['payment_type'] == "annual":
            data_goal.append([item['financial_goal'] , item['amount'] , item['amount'] , item['amount'] , item['amount'] , item['amount'] , item['amount']])
        if item['payment_type'] == "one-time":
            rowData = ["0"] * 7  
            rowData[0] = item['financial_goal']  
            print("item['time_period']: ", item['time_period'])
            if 1 <= int(item['time_period']) <= 6:
                rowData[int(item['time_period'])+1] = item['amount']  
            
            print("rowData: ", rowData)
            data_goal.append(rowData)
        
            
    # data = [
    #     ["1" , "0" , "0" , "0" , "0" , "0" , "0"],
    #     ["2" , "0" , "0" , "0" , "0" , "0" , "0"],
    #     ["3" , "0" , "0" , "0" , "0" , "0" , "0"],
    #     ["4" , "0" , "0" , "0" , "0" , "0" , "0"],
    #     ["5" , "0" , "0" , "0" , "0" , "0" , "0"],
    #     ["6" , "0" , "0" , "0" , "0" , "0" , "0"]
    # ]

    for i in data_goal:
        row = row + 1
        print("i :", i)
        ws.append(i)

    

    row = row + 1

    sum_columns = [sum(int(row[i]) for row in data_goal) for i in range(1, 7)]
    # แสดงผลลัพธ์
    print(sum_columns)


    ws.append(["รวมรายจ่ายเพื่อบรรลุเป้าหมายทางการเงิน" , "{:,.2f}".format(sum_columns[0]) , "{:,.2f}".format(sum_columns[1]) , "{:,.2f}".format(sum_columns[2]) , "{:,.2f}".format(sum_columns[3]) , "{:,.2f}".format(sum_columns[4]) , "{:,.2f}".format(sum_columns[5])]) 
    ws["A"+str(row)].font = Font(color="FF0000") 
    ws["B"+str(row)].font = Font(color="FF0000") 
    ws["C"+str(row)].font = Font(color="FF0000") 
    ws["D"+str(row)].font = Font(color="FF0000") 
    ws["E"+str(row)].font = Font(color="FF0000") 
    ws["F"+str(row)].font = Font(color="FF0000") 
    ws["G"+str(row)].font = Font(color="FF0000")

    row = row + 1
    
    ws.append(["กระแสเงินสดสุทธิคงเหลือหลังเป้าหมาย" , "{:,.2f}".format(((income_1 + sumother + income_4) - ((fix + insurance) + varSum))-sum_columns[0]) , "{:,.2f}".format(((income_1_2568 + sumother + income_4_2568) - ((fix_2568 + insurance_2568) + varSum_2568))-sum_columns[1]) , "{:,.2f}".format(((income_1_2569 + sumother + income_4_2569) - ((fix_2569 + insurance_2569) + varSum_2569))-sum_columns[2]) , "{:,.2f}".format(((income_1_2570 + sumother + income_4_2570) - ((fix_2570 + insurance_2570) + varSum_2570))-sum_columns[3]) , "{:,.2f}".format(((income_1_2571 + sumother + income_4_2571) - ((fix_2571 + insurance_2571) + varSum_2571))-sum_columns[4]) , "{:,.2f}".format(((income_1_2572 + sumother + income_4_2572) - ((fix_2572 + insurance_2572) + varSum_2572))-sum_columns[5])]) 
    ws["A"+str(row)].font = Font(color="FF0000") 
    ws["B"+str(row)].font = Font(color="FF0000") 
    ws["C"+str(row)].font = Font(color="FF0000") 
    ws["D"+str(row)].font = Font(color="FF0000") 
    ws["E"+str(row)].font = Font(color="FF0000") 
    ws["F"+str(row)].font = Font(color="FF0000") 
    ws["G"+str(row)].font = Font(color="FF0000")

    row = row + 1

    row_set_table_one = row
    ws.append(["กระแสเงินสดสุทธิคงเหลือหลังเป้าหมาย(ทบยอด)" , "{:,.2f}".format(sum_all -sum_columns[0]) , "{:,.2f}".format(sum_all_2668 -sum_columns[1]) , "{:,.2f}".format(sum_all_2669 -sum_columns[2])  , "{:,.2f}".format(sum_all_2670 -sum_columns[3])  , "{:,.2f}".format(sum_all_2671 -sum_columns[4])  , "{:,.2f}".format(sum_all_2672 -sum_columns[5]) ]) 
    ws["A"+str(row)].font = Font(color="FF0000") 
    ws["B"+str(row)].font = Font(color="FF0000") 
    ws["C"+str(row)].font = Font(color="FF0000") 
    ws["D"+str(row)].font = Font(color="FF0000") 
    ws["E"+str(row)].font = Font(color="FF0000") 
    ws["F"+str(row)].font = Font(color="FF0000") 
    ws["G"+str(row)].font = Font(color="FF0000")


    ## ----- ตาราง 2
    row = row + 2
    ws.append(["", "", ""])
    ws.append(["", "", ""])
    row_set_table = row

    row = row + 1
    ws.append(["วางแผนเกษียณ", ""])  
    ws.merge_cells("A"+str(row)+":A"+str(row+1))
    ws["A"+str(row)].alignment = Alignment(horizontal="center")
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = footheader_fill
    ws["A"+str(row)].font = Font(bold=True)

    ws["B"+str(row)] = "ปัจจุบัน"
    ws["B"+str(row)].alignment = Alignment(horizontal="center")
    ws["B"+str(row)].font = Font(bold=True)
    ws["B"+str(row)].fill = footheader_fill
    ws["B"+str(row)].font = Font(bold=True)

    ws["B"+str(row+1)] = "2567"
    ws["B"+str(row+1)].alignment = Alignment(horizontal="center")
    ws["B"+str(row+1)].font = Font(bold=True)
    ws["B"+str(row+1)].fill = footheader_fill
    ws["B"+str(row+1)].font = Font(bold=True)


    print("personal_info :", personal_info)
    print("financial_info: ", financial_info)

  
    yearly_expenses = expense_sum
    years_before_retirement = int(financial_info['retirement_age']) - int(personal_info['age'])
    fund_on_retirement = calculate_investment(
        principal=Decimal(yearly_expenses),
        annual_contribution=Decimal(0),
        annual_rate=3,  # Convert to percentage if applicable
        years=years_before_retirement
    )

    print("fund_on_retirement:" , fund_on_retirement)

    fix_percent = 5 / 100  #
    fix_data = 3 / 100  

    # คำนวณ
    result = ((1 + fix_percent) / (1 + fix_data)) - 1

    data = [
       ["อายุปัจจุบัน" , personal_info['age']],
       ["อายุที่คาดว่าจะเกษียณ" , financial_info['retirement_age']],
       ["อายุที่คาดว่าจะเสียชีวิต" , financial_info['life_expectancy']],
       ["รายจ่ายต่อปี" , "{:,.2f}".format(expense_sum)],
       ["เงินเฟ้อ" , "3%"],
       ["จำนวนเงินที่ต้องมี ณ วันแรกที่เกษียณ(รายจ่าย ณ วันเกษียณ)" , "{:,.2f}".format(fund_on_retirement)],
       ["อัตราผลตอบแทนหลังเกษียณ" , "%5"],
       ["Discount Rate" , "{:,.2f}".format(result * 100)],
    ]

    for i in data:
        row = row + 1
        ws.append(i)


    row = row + 2
    ## -------- ทำถึงอันนี้ -----------
    must_have_fund_on_retirment = calculate.must_have_fund_on_retirment(financial_info['retirement_age'] , financial_info['life_expectancy'] , fund_on_retirement , result)
    
    print("pv_calculate: ", must_have_fund_on_retirment)
    ws.append(["เงินที่ต้องมีไว้ใช้หลังเกษียณ", "{:,.2f}".format(must_have_fund_on_retirment)])  
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = PatternFill(start_color="fbe599", end_color="fbe599", fill_type="solid")
    ws["B"+str(row)].fill = PatternFill(start_color="fbe599", end_color="fbe599", fill_type="solid") 


    row = row + 1
    sum_asset = int(asset['number_bank']) + int(asset['fixedDeposit'])
    ws.append(["เงินออมปัจจุบัน", "{:,.2f}".format(sum_asset)])  
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = PatternFill(start_color="f4cccc", end_color="f4cccc", fill_type="solid") 
    ws["B"+str(row)].fill = PatternFill(start_color="f4cccc", end_color="f4cccc", fill_type="solid") 

    print("asset : ", asset)
    # sum_asset = int(asset['number_bank']) + int(asset['fixedDeposit'])

    income_1  =float(income['income_1']) if income['income_1'] != "" else 0
    income_2  =float(income['income_2']) if income['income_2'] != "" else 0
    income_3  =float(income['income_3']) if income['income_3'] != "" else 0
    income_4  =float(income['income_4']) if income['income_4'] != "" else 0
    income_5  =float(income['income_5']) if income['income_5'] != "" else 0
    income_6  =float(income['income_6']) if income['income_6'] != "" else 0
    income_7  =float(income['income_7']) if income['income_7'] != "" else 0
    income_8  =float(income['income_8']) if income['income_8'] != "" else 0
    
    sum_income = income_1 + income_2 +income_3 +income_4 +income_5 +income_6 +income_7 +income_8
    sumall = (sum_income - expense_sum)


    row = row + 1

    print(" --------- !!")

    will_have_fund_on_retirment = calculate.will_have_fund_on_retirment(sum_asset , sumall, financial_info['retirement_age'], personal_info['age'] )

    ws.append(["เงินที่จะมีตอนเกษียณ", "{:,.2f}".format(will_have_fund_on_retirment) ])  
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = PatternFill(start_color="fbe599", end_color="fbe599", fill_type="solid")
    ws["B"+str(row)].fill = PatternFill(start_color="fbe599", end_color="fbe599", fill_type="solid")  

    # missing_fund = will_have_fund_on_retirment - must_have_fund_on_retirment
    # missing_PMT = ((missing_fund*THOR/100)/(((1+THOR/100)**(retirement_age-int(personal_info['age'])))-1))- total_cashflow
    # if missing_fund > 0:
    #     print("เงินเพียงพอสำหรับการเกษียณ")#ขาดไป
    #     ws.append(["ขาดไป", "เงินเพียงพอสำหรับการเกษียณ"]) 
    # else:
    #     print(missing_fund)#ขาดไป
    #     ws.append(["ขาดไป", "{:,.2f}".format(missing_fund) ]) 
    
    # print(missing_PMT)#ต้องออมเพิ่มขึ้นเดือนละ
    print(" --------- ")
    missing_PMT , message = calculate.sum_have_fund_on_retirment(will_have_fund_on_retirment, must_have_fund_on_retirment , financial_info['retirement_age'], personal_info['age'] , sumall)

    row = row + 1
    ws.append(["ขาดไป", message ]) 
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = PatternFill(start_color="f4cccc", end_color="f4cccc", fill_type="solid") 
    ws["B"+str(row)].fill = PatternFill(start_color="f4cccc", end_color="f4cccc", fill_type="solid") 

    row = row + 1
    row_stop_table = row
    
    ws.append(["ต้องออมเพิ่มขึ้นปีละ", "{:,.2f}".format(missing_PMT)])  
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = PatternFill(start_color="f4cccc", end_color="f4cccc", fill_type="solid") 
    ws["B"+str(row)].fill = PatternFill(start_color="f4cccc", end_color="f4cccc", fill_type="solid")   


    for row in ws.iter_rows(min_row=1, max_row=row_set_table_one, min_col=1, max_col=7):
        if not all(cell.value == "" for cell in row):
            for cell in row:
                cell.border = thin_border


    for row in ws.iter_rows(min_row=(row_set_table+1), max_row=row_stop_table, min_col=1, max_col=2):  # ตารางสอง 2 คอลัมน์ (A:B)
        for cell in row:
            cell.border = thin_border
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    
    wb.save(filename)
    print(f"สร้างไฟล์ {filename} สำเร็จ!")
    return True




def create_budget_excel(asset , asset_detail , liabilities , income , expense_detail , filename="budget.xlsx"):
    row = 1
    wb = Workbook()
    ws = wb.active
    ws.title = "งบการเงิน"
    
    header_fill = PatternFill(start_color="d5a6bd", end_color="d5a6bd", fill_type="solid")  
    subheader_fill = PatternFill(start_color="f4cccc", end_color="f4cccc", fill_type="solid") 
    footheader_fill = PatternFill(start_color="fbe599", end_color="fbe599", fill_type="solid") 
    
    thin_border = Border(left=Side(style='thin', color="000000"),
                         right=Side(style='thin', color="000000"),
                         top=Side(style='thin', color="000000"),
                         bottom=Side(style='thin', color="000000"))
    bold_border = Border(left=Side(style='medium', color="000000"),
                         right=Side(style='medium', color="000000"),
                         top=Side(style='medium', color="000000"),
                         bottom=Side(style='medium', color="000000"))
    
    asset_sum = 0

    ws.append(["งบสถานะบุคคล", "", ""])  
    ws.merge_cells("A"+str(row)+":C"+str(row))
    ws["A"+str(row)].alignment = Alignment(horizontal="center")
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = header_fill
    ws["A"+str(row)].font = Font(bold=True)
    
    row = row + 1
    ws.append(["สินทรัพย์", "", "%"])
    ws.merge_cells("A"+str(row)+":B"+str(row))
    ws["A"+str(row)].alignment = Alignment(horizontal="center")
    ws["A"+str(row)].fill = header_fill
    ws["B"+str(row)].fill = header_fill
    ws["C"+str(row)].fill = header_fill

    row = row + 1
    ws.append(["สินทรัพย์สภาพคล่อง", "", ""])
    ws.merge_cells("A"+str(row)+":C"+str(row))
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = subheader_fill
    ws["C"+str(row)].fill = subheader_fill

    asset_sum_total = 0
    if asset['number_bank'] != "" and asset['fixedDeposit'] != "":
        asset_sum_total = asset_sum_total + (int(asset['number_bank']) + int(asset['fixedDeposit']))
    
    if len(asset_detail) > 0: 
        for item in asset_detail:
            print("item :" , item)
            asset_sum_total = asset_sum_total + int(item['asset_value'])

    print("asset_sum_total :" , asset_sum_total)
    number_bank = (float(asset['number_bank']) / asset_sum_total) * 100
    fixedDeposit = (float(asset['fixedDeposit']) / asset_sum_total) * 100
    asset_data = [
        ["เงินสด", int(asset['number_bank']),  "{:,.2f}%".format(number_bank)  ],
        ["บัญชีฝากออมทรัพย์", int(asset['fixedDeposit']), "{:,.2f}%".format(fixedDeposit) ],
    ]

    for rowData in asset_data:
        row = row + 1
        ws.append(rowData)

    row = row + 1
    print("asset : ", asset)
    sumNumber_bankAndfixedDeposit = (int(asset['number_bank']) + int(asset['fixedDeposit']))
    print("-------")
    fixedDepositSum = ( float(sumNumber_bankAndfixedDeposit) / asset_sum_total) * 100
    ws.append(["รวมสินทรัพย์สภาพคล่อง", sumNumber_bankAndfixedDeposit,  "{:,.2f}%".format(fixedDepositSum)])
    ws["A"+str(row)].font = Font(bold=True)

    asset_sum = asset_sum + (int(asset['number_bank']) + int(asset['fixedDeposit']))
    row = row + 1
    ws.append(["สินทรัพย์เพื่อการลงทุน", "", ""])
    ws.merge_cells("A"+str(row)+":C"+str(row))
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = subheader_fill
    ws["C"+str(row)].fill = subheader_fill
    
    # row = row + 1
    # ws.append(["หุ้นสามัญ", "", ""])
    # ws["A"+str(row)].font = Font(bold=True)
    type_mapping = {
        "common_shares": "หุ้นสามัญ",
        "funds": "กองทุน",
        "bonds": "พันธบัตร/หุ้นกู้",
        "perferred_shares": "หุ้นบุริมสิทธิ",
        "personal_asset": "สินทรัพย์ส่วนบุคคล",
        "investment": "สินทรัพย์เพื่อการลงทุนอื่น ๆ",
        "other": "ทรัพย์สินอื่น ๆ"
    }

    print("-------!!!")
    grouped_assets = defaultdict(list)

    print("asset_detail :" , asset_detail)
    if len(asset_detail) > 0:
        for item in asset_detail:
            asset_type = type_mapping.get(item['asset_type'], item['asset_type'])
            grouped_assets[asset_type].append(item)

    asset_data = []

    print("-------!!!")
    print("grouped_assets: ",  grouped_assets)
    
    if len(grouped_assets) > 0:
        for asset_type, assets in grouped_assets.items():
            asset_data.append([asset_type, "", ""])
            for asset in assets:
                profit = asset['asset_value']
                print("profit :" , profit)
                asset_sum = asset_sum + int(profit)
                sumdata = ( float(profit) / asset_sum_total) * 100
                asset_data.append([asset['asset_name'], int(profit), "{:,.2f}%".format(sumdata)])

    print("asset_data: " , asset_data)
    if len(asset_data) > 0:
        for rowData in asset_data:
            row = row + 1
            print("rowData: ", rowData)
            if rowData[0] in ["สินทรัพย์ส่วนบุคคล", "สินทรัพย์เพื่อการลงทุนอื่น ๆ", "ทรัพย์สินอื่น ๆ"]:
                ws.append(rowData)
                ws.merge_cells("A"+str(row)+":C"+str(row))
                ws["A"+str(row)].font = Font(bold=True)
                ws["A"+str(row)].fill = subheader_fill
                ws["C"+str(row)].fill = subheader_fill
            else:
                ws.append(rowData)

    row = row + 1
    print("!!!----------_!!!")
    sumdataall = ( float(asset_sum_total) / asset_sum_total) * 100
    ws.append(["สินทรัพย์รวม", int(asset_sum_total), "{:,.2f}%".format(sumdataall)])
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = header_fill
    ws["B"+str(row)].fill = header_fill
    ws["C"+str(row)].fill = header_fill
    
    print("!!!----------_!!!----")
    ## ----- ตารางที่ 2 ขขขขขขข
    row = row + 3
    ws.append(["", "", ""])
    ws.append(["", "", ""])
    ws.append(["", "", ""])
    
    row = row + 1
    ws.append(["หนี้สิน", "", "%"])
    ws.merge_cells("A"+str(row)+":B"+str(row))
    ws["A"+str(row)].alignment = Alignment(horizontal="center")
    ws["A"+str(row)].fill = header_fill
    ws["B"+str(row)].fill = header_fill
    ws["C"+str(row)].fill = header_fill
    # ws[ws.max_row][0].fill = header_fill

    row = row + 1
    ws.append(["หนี้สินระยะสั้น", "", ""])
    ws.merge_cells("A"+str(row)+":C"+str(row))
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = subheader_fill
    ws["C"+str(row)].fill = subheader_fill

    liability_data_short_trem = []
    liability_data_long_trem = []
    liability_sum_short_trem = 0
    liability_sum_long_trem = 0

    # (asset_sum_total - liability_sum) + liability_sum
    if len(liabilities) > 0:
        for item in liabilities:
            if item['liability_type'] == "short-term":
                liability_sum_short_trem = liability_sum_short_trem +  int(item['amount_due'])
            if item['liability_type'] == "long-term":
                liability_sum_long_trem = liability_sum_long_trem +  int(item['amount_due'])

    liability_sum = liability_sum_long_trem + liability_sum_short_trem
    liability_sum_all =  (asset_sum_total - liability_sum) + liability_sum

    if len(liabilities) > 0:
        for item in liabilities:
            if item['liability_type'] == "short-term":
                sum_liability_sum_all = ( int(item['amount_due']) / float(liability_sum)) * 100
                liability_data_short_trem.append([check_type_short_term(item['liability_name']) , item['amount_due'], "{:,.2f}%".format(sum_liability_sum_all)])
            if item['liability_type'] == "long-term":
                sum_liability_sum_all = ( int(item['amount_due']) / float(liability_sum)) * 100
                liability_data_long_trem.append([check_type_long_term(item['liability_name']) , item['amount_due'], "{:,.2f}%".format(sum_liability_sum_all)])
    

    print("liability_data_short_trem " , liability_data_short_trem)
    if len(liability_data_short_trem) > 0:
        for rowData in liability_data_short_trem:
            row = row + 1
            ws.append(rowData)

    row = row + 1
    sum_liability_sum_all = 0
    print("liability_sum_short_trem :" ,liability_sum_short_trem)
    print("liability_sum :" ,liability_sum)
    if liability_sum != 0:
        print("sum_liability_sum_all > ")
        sum_liability_sum_all = ( int(liability_sum_short_trem) / float(liability_sum)) * 100
    print("sum_liability_sum_all :" ,sum_liability_sum_all)
    ws.append(["รวมหนี้สินระยะสั้น", int(liability_sum_short_trem), "{:,.2f}%".format(sum_liability_sum_all)])
    ws["A"+str(row)].font = Font(bold=True)

    row = row + 1
    ws.append(["หนี้สินระยะยาว", "", ""])
    ws.merge_cells("A"+str(row)+":C"+str(row))
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = subheader_fill
    ws["C"+str(row)].fill = subheader_fill

    # liability_data = [
    #     ["1", "#REF!", "#REF!"],
    # ]
    print("liability_data_long_trem :" , liability_data_long_trem)
    if len(liability_data_long_trem) > 0:
        for rowData in liability_data_long_trem:
            row = row + 1
            ws.append(rowData)

    row = row + 1
    sum_liability_sum_all = 0
    if liability_sum != 0:
        sum_liability_sum_all = ( int(liability_sum_long_trem) / float(liability_sum)) * 100
    ws.append(["รวมหนี้สินระยะยาว", int(liability_sum_long_trem) , "{:,.2f}%".format(sum_liability_sum_all)])
    ws["A"+str(row)].font = Font(bold=True)

    row = row + 1
    sum_liability_sum_all = 0
    if liability_sum != 0:
        sum_liability_sum_all = ( int(liability_sum) / float(liability_sum)) * 100
    ws.append(["หนี้สินรวม", liability_sum_long_trem + liability_sum_short_trem , "{:,.2f}%".format(sum_liability_sum_all)])
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = header_fill
    ws["B"+str(row)].fill = header_fill
    ws["C"+str(row)].fill = header_fill

    print("!!!!-------------")

    row = row + 1
    ws.append(["ความมั่งคั่งสุทธิ", asset_sum_total - liability_sum, ""])
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = footheader_fill
    ws["B"+str(row)].fill = footheader_fill
    ws["C"+str(row)].fill = footheader_fill


    row = row + 1
    ws.append(["หนี้สินและความมั่งคั่งสุทธิ", liability_sum_all , ""])
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = footheader_fill
    ws["B"+str(row)].fill = footheader_fill
    ws["C"+str(row)].fill = footheader_fill


     ## ----- ตารางที่ 3 ขขขขขขข
    row = row + 3
    ws.append(["", "", ""])
    ws.append(["", "", ""])
    ws.append(["", "", ""])

    row = row + 1
    ws.append(["งบกระแสเงินสด", "", ""])  
    ws.merge_cells("A"+str(row)+":C"+str(row))
    ws["A"+str(row)].alignment = Alignment(horizontal="center")
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = header_fill
    ws["A"+str(row)].font = Font(bold=True)

    row = row + 1
    ws.append(["กระแสเงินสดรับ", "", "%"])
    ws.merge_cells("A"+str(row)+":B"+str(row))
    ws["A"+str(row)].alignment = Alignment(horizontal="center")
    ws["A"+str(row)].fill = header_fill
    ws["B"+str(row)].fill = header_fill
    ws["C"+str(row)].fill = header_fill
    
    # {
    #     'job': 'รับจ้างทั่วไป', 
    #     'work_place': 'ธนาคารกรุงเทพ สำนักงานใหญ่', 
    #     'department': 'ฝ่ายขาย', 'job_title': 'พนักงานขายกองทุนรวม', 
    #     'salary': '35000', 'bonus': '3', 
    #     'salary_growth': '6', 'working_years': '7', 
    #     'job_change_prop': '10'}
    print("income  :" , income)

    income_1  =float(income['income_1']) if income['income_1'] != "" else 0
    income_2  =float(income['income_2']) if income['income_2'] != "" else 0
    income_3  =float(income['income_3']) if income['income_3'] != "" else 0
    income_4  =float(income['income_4']) if income['income_4'] != "" else 0
    income_5  =float(income['income_5']) if income['income_5'] != "" else 0
    income_6  =float(income['income_6']) if income['income_6'] != "" else 0
    income_7  =float(income['income_7']) if income['income_7'] != "" else 0
    income_8  =float(income['income_8']) if income['income_8'] != "" else 0
    print("income_1 --- + :" , income_1)
    sum_income = income_1 + income_2 +income_3 +income_4 +income_5 +income_6 +income_7 +income_8
    print("sum_income --- :" , sum_income)
    
    percent_income_1 = ( float(income_1) / float(sum_income)) * 100 if income_1 != 0 else 0.0
    percent_income_2 = ( float(income_2) / float(sum_income)) * 100 if income_2 != 0 else 0.0
    percent_income_3 = ( float(income_3) / float(sum_income)) * 100 if income_3 != 0 else 0.0
    percent_income_4 = ( float(income_4) / float(sum_income)) * 100 if income_4 != 0 else 0.0
    percent_income_5 = ( float(income_5) / float(sum_income)) * 100 if income_5 != 0 else 0.0
    percent_income_6 = ( float(income_6) / float(sum_income)) * 100 if income_6 != 0 else 0.0
    percent_income_7 = ( float(income_7) / float(sum_income)) * 100 if income_7 != 0 else 0.0
    percent_income_8 = ( float(income_8) / float(sum_income)) * 100 if income_8 != 0 else 0.0

    data = [
        ["มาตรา 40(1) งานประจำ", income_1 , "{:,.2f}%".format(percent_income_1) ],
        ["มาตรา 40(2) ค่าจ้างทั่วไป", income_2 , "{:,.2f}%".format(percent_income_2)],
        ["มาตรา 40(3) ค่าสิขสิทธิ์ สิทธิ์ในทรัพย์สินทางปัญญา", income_3 , "{:,.2f}%".format(percent_income_3)],
        ["มาตรา 40(4) ดอกเบี้ย&เงินปันผล", income_4 , "{:,.2f}%".format(percent_income_4)],
        ["มาตรา 40(5) ค่าเช่า เช่น เช่าบ้าน, เช่าพาหนะ", income_5 , "{:,.2f}%".format(percent_income_5)],
        ["มาตรา 40(6) ค่าวิชาชีพอิสระ", income_6 , "{:,.2f}%".format(percent_income_6)],
        ["มาตรา 40(7) ค่ารับเหมา", income_7 , "{:,.2f}%".format(percent_income_7)],
        ["มาตรา 40(8) เงินได้อื่นๆ เช่น กิจการซื้อมาขายไป", income_8 , "{:,.2f}%".format(percent_income_8)],
    ]

    print("data :" , data)
    for rowData in data:
        row = row + 1
        ws.append(rowData)

    row = row + 1
    print("sum_income " , sum_income)
    sum_income_all = 0.0
    if sum_income != 0:
        sum_income_all = ( int(sum_income) / float(sum_income)) * 100
    ws.append(["กระแสเงินสดรับรวม", sum_income , "{:,.2f}%".format(sum_income_all)])
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = header_fill
    ws["B"+str(row)].fill = header_fill
    ws["C"+str(row)].fill = header_fill

    expense_fix = []
    expense_insurance = []
    expense_var = []
    expense_sum = 0
    sum_expense_all = 0
    print("expense_detail: " , expense_detail)
    for item in expense_detail:
        if item['fix_cf_value'] != "":
            print("item['fix_cf_value'] : ", item['fix_cf_value'])
            print("sum_income : ", sum_income)
            if sum_income != 0:
                sum_expense_all = ( int(item['fix_cf_value']) / int(sum_income)) * 100
            
            if expense_sum != 0:
                expense_sum = expense_sum + int(item['fix_cf_value'])
        if item['fix_cf_mode'] == "fix":
            if item['fix_cf_value'] != "":
                expense_fix.append([item['fix_cf_name'], item['fix_cf_value'], "{:,.2f}%".format(sum_expense_all)])
        if item['fix_cf_mode'] == "insurance":
            if item['fix_cf_value'] != "":
                expense_insurance.append([ check_type_expense_insurance(item['fix_cf_name']), item['fix_cf_value'], "{:,.2f}%".format(sum_expense_all)])
        if item['fix_cf_mode'] == "var":
            if item['fix_cf_value'] != "":
                expense_var.append([check_type_expense_var(item['fix_cf_name']), item['fix_cf_value'], "{:,.2f}%".format(sum_expense_all)])
        print("expense_detail : " , item)

    row = row + 1
    ws.append(["เงินงวดผ่อนชําระคืนหนี้", "", ""])
    ws["A"+str(row)].font = Font(bold=True)

    

    # data = [
    #     ["เงินกู้ซื้อรถยนต์", "0", "0%"],
    #     ["บัตรเครดิต", "0", "0%"],
    # ]
    if len(expense_fix) > 0:
        for rowData in expense_fix:
            row = row + 1
            ws.append(rowData)

    row = row + 1
    ws.append(["เบี้ยประกัน", "", ""])
    ws["A"+str(row)].font = Font(bold=True)

    # data = [
    #     ["รถยนต์", "0", "0%"],
    #     ["ประกันสังคม", "0", "0%"],
    #     ["เงินสะสมกองทุนสำรองเลี้ยงชีพ", "0", "0%"],
    # ]
    print("expense_insurance :" , expense_insurance)
    if len(expense_insurance) > 0:
        for rowData in expense_insurance:
            row = row + 1
            ws.append(rowData)

    row = row + 1
    ws.append(["กระแสเงินสดจ่ายผันแปร", "", ""])
    ws["A"+str(row)].font = Font(bold=True)

    # data = [
    #     ["ค่าอาหาร", "0", "0%"],
    #     ["ค่าโทรศัพท์", "0", "0%"],
    #     ["ค่าใช้จ่ายนันทนาการ", "0", "0%"],
    # ]
    if len(expense_var) > 0:
        for rowData in expense_var:
            row = row + 1
            ws.append(rowData)

    row = row + 1
    print("expense_sum : ", expense_sum)
    sum_expense_all = 0
    if expense_sum != 0:
        sum_expense_all = ( int(expense_sum) / float(sum_income)) * 100
    ws.append(["กระแสเงินสดจ่ายรวม", expense_sum , "{:,.2f}%".format(sum_expense_all)])
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].alignment = Alignment(horizontal="center")
    ws["A"+str(row)].fill = header_fill
    ws["B"+str(row)].fill = header_fill
    ws["C"+str(row)].fill = header_fill


    row = row + 1
    print("sum_income :" , sum_income)
    print("expense_sum :" , expense_sum)
    sumall = 0
    sum_expense_all = 0.0
    if expense_sum != 0:
        sumall = (sum_income - expense_sum)
        sum_expense_all = ( int(sumall) / float(sum_income)) * 100
    ws.append(["กระแสเงินสดสุทธิ", sumall, "{:,.2f}%".format(sum_expense_all)])
    ws["A"+str(row)].font = Font(bold=True)
    ws["A"+str(row)].fill = footheader_fill
    ws["B"+str(row)].fill = footheader_fill
    ws["C"+str(row)].fill = footheader_fill
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        if not all(cell.value == "" for cell in row):
            for cell in row:
                cell.border = thin_border
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    
    wb.save(filename)
    print(f"สร้างไฟล์ {filename} สำเร็จ!")
    return True

# create_cash_flow_excel("cash_flow.xlsx")
